from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, authenticate, logout
from django.http import JsonResponse
from django.contrib import messages
from django.db.models import Sum, Q
from django.conf import settings
from .models import Employee, Advance, Deduction, Arrears
from .forms import EmployeeForm, AdvanceForm, DeductionForm, ArrearsForm
from django.core.paginator import Paginator
from django.utils import timezone
from datetime import timedelta, datetime
from django.http import HttpResponse
from django.template.loader import render_to_string
from xhtml2pdf import pisa
from openpyxl import Workbook
import base64, os
# Create your views here.

def index(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password= request.POST.get('password')
        
        user = authenticate(username=username,password=password)
        
        if user is not None:
            login(request,user)
            return redirect('dashboard')
        
        else:
            messages.error(request, 'Invalid Credentials')
            return redirect('index')            

    else:
        return render(request, 'index.html')

def dashboard(request):
    employees = Employee.objects.all()
    advances = Advance.objects.all()
    employee_count = Employee.objects.count()
    payroll_amount = Employee.objects.aggregate(Sum('salary'))['salary__sum'] or 0
    advance_amounts= Advance.objects.aggregate(Sum('amount'))['amount__sum'] or 0
    netpay_total= sum(employee.net_pay for employee in employees)
    
    recent_employees = Employee.objects.order_by('-employee_id')[:2]  # Last 5 employees added
    recent_advances = Advance.objects.order_by('-date')[:3]  # Last 5 advances
    upcoming_deadline = timezone.now() + timedelta(days=7)   # Example of a payroll deadline in 7 days
    
    # Format recent activities to show in the template
    recent_activities = [
        f"New employee added: {employee.first_name} {employee.last_name}" for employee in recent_employees
    ] + [
        f"Advance recorded: {advance.amount} on {advance.date}" for advance in recent_advances
    ] + [
        f"Upcoming payroll deadline: {upcoming_deadline.strftime('%b %d')}"
    ]
    
    chart_data = {
        'labels': ['Gross Salary', 'Advances', 'Net Pay'],
        'values': [
            float(employees.aggregate(Sum('salary'))['salary__sum'] or 0),
            float(advances.aggregate(Sum('amount'))['amount__sum'] or 0),
            float(sum(employee.net_pay for employee in employees)),
        ]
    }
        
    return render(request, 'dashboard.html', {'employee_count': employee_count, 'payroll_amount' : payroll_amount, 'advance_amounts':advance_amounts, 'netpay_total': netpay_total,'recent_activities':recent_activities,'chart_data':chart_data})

def employees(request):
    if request.method == 'POST':
        form = EmployeeForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('employees')
    else:
        form = EmployeeForm()
        
    query = request.GET.get('q')
    if query:
        employees = Employee.objects.filter(
            Q(last_name__icontains=query) | Q(first_name__icontains=query)
        )
    else:
        employees = Employee.objects.all()
    
    # Pagination
    paginator = Paginator(employees, 25)  # 25 employees per page
    page_number = request.GET.get('page')
    employees_page = paginator.get_page(page_number)

    return render(request, 'employees.html', {'form': form, 'employees': employees_page, 'query': query})

def generate_employee_list(request):
    employees= Employee.objects.all()
    
    wb= Workbook()
    ws= wb.active
    ws.title='Employees List'
    
    headers = [
        'First Name','Last Name','Phone Number','Gender','Designation','ID Number',
    ]
    
    ws.append(headers)
    
    for employee in employees:
        
        ws.append((
            employee.first_name,
            employee.last_name,
            employee.phone_number,
            employee.gender,
            employee.area_of_work,
            employee.id_number            
        ))
    
    response= HttpResponse(content_type= 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition']= 'attachment; filename="SPS Employee List.xlsx"'
    wb.save(response)
    
    return response
    
def get_employee_data(request, employee_id):
    employee = get_object_or_404(Employee, employee_id=employee_id)
    data = {
        'last_name': employee.last_name,
        'first_name': employee.first_name,
        'phone_number': employee.phone_number,
        'gender': employee.gender,
        'area_of_work': employee.area_of_work,
        'id_number': employee.id_number,
        'salary': str(employee.salary),
        'days_worked': employee.days_worked,
    }
    return JsonResponse(data)

def update_employee(request, employee_id):
    if request.method == "POST":
        employee = get_object_or_404(Employee, employee_id=employee_id)
        employee.last_name = request.POST.get('last_name')
        employee.first_name = request.POST.get('first_name')
        employee.phone_number = request.POST.get('phone_number')
        employee.gender = request.POST.get('gender')
        employee.area_of_work = request.POST.get('area_of_work')
        employee.id_number = request.POST.get('id_number')
        employee.salary = request.POST.get('salary')
        employee.days_worked = request.POST.get('days_worked')
        employee.save()
        return redirect('employees')

def delete_employee(request,employee_id):
    employee = get_object_or_404(Employee,employee_id=employee_id)
    employee.delete()
    return redirect('employees')

def advance(request):    
    if request.method == 'POST':
        form = AdvanceForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('advance')
            
    else:
        form= AdvanceForm()
        
    advances= Advance.objects.all()       
        
    return render(request, 'advance.html',{'advances':advances, 'form' : form})

def get_advance_data(request,advance_id):
    advance = get_object_or_404(Advance, id=advance_id)
    data = {
        'amount': advance.amount,
        'date': advance.date, 
    }
    
    return JsonResponse(data)

def update_advance(request,advance_id):
    if request.method == 'POST':
        advance= get_object_or_404(Advance,id=advance_id)
        advance.amount = request.POST.get('amount')
        advance.date = request.POST.get('date')
        
        advance.save()
        return redirect('advance')

def delete_advance(request,advance_id):
    advance= get_object_or_404(Advance,id=advance_id)
    advance.delete()
    return redirect('advance')

def deductions(request):
    if request.method == 'POST':
        form = DeductionForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('deductions')
    
    else:
        form = DeductionForm()
        
    deductions = Deduction.objects.all()
    
    return render(request, 'deductions.html', {'deductions': deductions, 'form' : form})

def get_deduction_data(request,deduction_id):
    deduction = get_object_or_404(Deduction,id=deduction_id)
    
    data = {
        'nhif' : deduction.nhif,
        'nssf' : deduction.nssf,
        'uniform': deduction.uniform,
        'fuel': deduction.fuel
    }
    
    return JsonResponse(data)

def update_deduction(request,deduction_id):
    if request.method == 'POST':
        deduction = get_object_or_404(Deduction,id=deduction_id)
        deduction.nhif = request.POST.get('nhif')
        deduction.nssf = request.POST.get('nssf')
        deduction.uniform = request.POST.get('uniform')
        deduction.fuel = request.POST.get('fuel')
        
        deduction.save()
        return redirect('deductions')
       
def delete_deductions(request,deduction_id):
    deduction=get_object_or_404(Deduction,id=deduction_id)
    deduction.delete()
    return redirect('deductions')

def log_out(request):
    logout(request)
    return redirect('index')

def arrears(request):
    if request.method == 'POST':
        form= ArrearsForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('arrears')
            
    else:
        form= ArrearsForm()
        
    arrears= Arrears.objects.all()
    
    return render(request,'arrears.html', {'form':form, 'arrears':arrears})

def get_arrears_data(request,arrears_id):
    arrears= get_object_or_404(Arrears, id=arrears_id)
   
    data = {
        'amount' : arrears.amount,
        'month' : arrears.month,
    }
    
    return JsonResponse(data)

def update_arrears(request,arrears_id):
    if request.method == 'POST':
        arrears =get_object_or_404(Arrears,id=arrears_id)
        
        arrears.amount= request.POST.get('amount'),
        arrears.month = request.POST.get('month')
        
        arrears.save()
        
        return redirect('arrears')

def delete_arrears(request,arrears_id):
    arrears = get_object_or_404(Arrears,id=arrears_id)
    arrears.delete()
    return redirect('arrears')

def generate_payslip(request,employee_id):
    employee = Employee.objects.get(employee_id=employee_id)

    net_pay = employee.net_pay
    deductions_total = Deduction.objects.filter(employee=employee).aggregate(
        nhif=Sum('nhif') or 0,
        nssf=Sum('nssf') or 0,
        uniform=Sum('uniform') or 0,
        fuel=Sum('fuel') or 0
    )
    
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'logo.jpg')
    with open(logo_path, 'rb') as logo_file:
        logo_base64 = base64.b64encode(logo_file.read()).decode('utf-8')
    
    payslip_data = {
        'employee' : employee,
        'net_pay' : net_pay,
        'advances' : Advance.objects.filter(employee=employee),
        'deductions' : Deduction.objects.filter(employee=employee),
        'arrears' : Arrears.objects.filter(employee=employee),
        'date' : timezone.now().strftime('%Y-%m-%d'),
        'deductions_sum': sum(value or 0 for value in deductions_total.values()),
        'logo_path': logo_path
    }
    
    html_content = render_to_string('payslip.html',payslip_data)
    
    response = HttpResponse(content_type='application/pdf')
    response["Content-Disposition"] = f'inline; filename="payslip_{employee.employee_id}.pdf"'

    
    pisa_status = pisa.CreatePDF(html_content, dest=response)
    
    if pisa_status.err:
        print('An Error Occurred While Generating Payslip', status=500)
        
    return response

def payroll_report(request):
    employees = Employee.objects.all()
    
    # Prepare data and calculate totals
    total_advances = 0
    total_deductions = 0
    total_arrears = 0
    total_net_pay = 0
    total_employees = employees.count()

    report_data = []
    for employee in employees:
        advances_total = Advance.objects.filter(employee=employee).aggregate(Sum('amount'))['amount__sum'] or 0
        arrears_total = Arrears.objects.filter(employee=employee).aggregate(Sum('amount'))['amount__sum'] or 0
        deductions_total = Deduction.objects.filter(employee=employee).aggregate(
            nhif=Sum('nhif') or 0,
            nssf=Sum('nssf') or 0,
            uniform=Sum('uniform') or 0,
            fuel=Sum('fuel') or 0
        )
        deductions_sum = sum(value or 0 for value in deductions_total.values())
        net_pay = employee.salary - advances_total - deductions_sum + arrears_total

        # Update totals
        total_advances += advances_total
        total_deductions += deductions_sum
        total_arrears += arrears_total
        total_net_pay += net_pay

        report_data.append({
            'employee': employee,
            'advances_total': advances_total,
            'arrears_total': arrears_total,
            'deductions_sum': deductions_sum,
            'net_pay': net_pay,
        })

    context = {
        'report_data': report_data,
        'total_employees': total_employees,
        'total_advances': total_advances,
        'total_deductions': total_deductions,
        'total_arrears': total_arrears,
        'total_net_pay': total_net_pay,
    }

    return render(request, 'report.html', context)

def excel_report(request):
    employees= Employee.objects.all()
    
    wb= Workbook()
    payroll_ws= wb.active
    payroll_ws.title="Payroll Report"
    advances_ws= wb.create_sheet(title='Advances')
    deductions_ws= wb.create_sheet(title='Deductions')
    arrears_ws=wb.create_sheet(title='Arrears')
    
    payroll_headers= [
        "Employee Name","Phone Number","Designation","Days Worked","Gross Pay","Advances","Deduction","Arreas","Net Pay"
    ]
  
    payroll_ws.append(payroll_headers)
    
    for employee in employees:
        advances_total= Advance.objects.filter(employee=employee).aggregate(Sum('amount'))['amount__sum'] or 0
        arrears_total= Arrears.objects.filter(employee=employee).aggregate(Sum('amount'))['amount__sum'] or 0
        deductions_total = Deduction.objects.filter(employee=employee).aggregate(
            nhif= Sum('nhif') or 0,
            nssf= Sum('nssf') or 0,
            uniform= Sum('uniform') or 0,
            fuel= Sum('fuel') or 0
        )
        deductions_sum = sum(value or 0 for value in deductions_total.values())
        net_pay= employee.salary - advances_total - deductions_sum + arrears_total
        
        payroll_ws.append([
        f"{employee.first_name} {employee.last_name}",
        employee.phone_number,
        employee.area_of_work,
        employee.days_worked,
        employee.salary,
        advances_total,
        deductions_sum,
        arrears_total,
        net_pay
    ])
    
    total_advances = Advance.objects.aggregate(Sum('amount'))['amount__sum'] or 0
    total_arrears = Arrears.objects.aggregate(Sum('amount'))['amount__sum'] or 0
    total_deductions= sum(
        value or 0 for value in Deduction.objects.aggregate(
            nhif= Sum('nhif') or 0,
            nssf= Sum('nssf') or 0,
            uniform= Sum('uniform') or 0,
            fuel= Sum('fuel') or 0,
        ).values()
    )
    total_net_pay= sum(emp.net_pay for emp in employees)
    
    payroll_ws.append([])
    payroll_ws.append(['TOTAL','','','','',total_advances,total_deductions,total_arrears,total_net_pay])
    
    
    
    advances_headers = [
        'Employee Name','amount','date'
    ]
    
    advances_ws.append(advances_headers)
    
    for advances in Advance.objects.all():
        advances_ws.append([
            f'{advances.employee.first_name} {advances.employee.last_name}',
            advances.amount,
            advances.date,
        ])
    
    
    deductions_headers = [
        'Employee Name','NHIF','NSSF','Uniform','Fuel'
    ]
    
    deductions_ws.append(deductions_headers)
    
    for deductions in Deduction.objects.all():
        deductions_ws.append([
            f'{deductions.employee.first_name} {deductions.employee.last_name}',
            deductions.nhif or 0,
            deductions.nssf or 0,
            deductions.uniform or 0,
            deductions.fuel or 0,
        ])
        
    
    arrears_headers = [
        'Employee','amount','month'
    ]
    
    arrears_ws.append(arrears_headers)
    
    for arrears in Arrears.objects.all():
        arrears_ws.append([
            f'{arrears.employee.first_name} {arrears.employee.last_name}',
            arrears.amount,
            arrears.month
        ])
     
        
    response= HttpResponse(content_type= 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition']= 'attachment; filename="SPS Payroll Report.xlsx"'
    wb.save(response)
    
    return response

